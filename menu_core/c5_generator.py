import os
import glob
import pandas as pd
import openpyxl

def generate_c5_from_dir(output_dir: str, raw_outlet: str, applicator: str):
    """
    Generate C5 Excel file from extracted CSVs in output_dir.
    """
    template_xlsx = os.path.join(os.path.dirname(os.path.dirname(__file__)), "O. C5.xlsx")
    if not os.path.exists(template_xlsx):
        print(f"  [C5 Generator] Template {template_xlsx} tidak ditemukan.")
        return False

    # Read from raw menu Excel instead of CSV since GoFood extractor saves to Excel
    raw_xlsxs = glob.glob(os.path.join(output_dir, "*menu*.xlsx"))
    raw_xlsxs = [f for f in raw_xlsxs if not os.path.basename(f).startswith("C5")]

    if not raw_xlsxs:
        print(f"  [C5 Generator] File Excel mentah (raw menu) tidak ditemukan di {output_dir}")
        return False

    raw_xlsx = raw_xlsxs[0]

    try:
        df_items = pd.read_excel(raw_xlsx, sheet_name='Items')
        df_modifiers = pd.read_excel(raw_xlsx, sheet_name='Modifiers')

        # Process Items
        items_out = pd.DataFrame(columns=[
            'OFD', 'Outlet Name', 'Outlet Short Name', 'Outlet Link', 'SID', 'Category ID', 'Category', 'Item ID', 'Item',
            'Photo Link', 'Description', 'Total Sold', 'Total Modifier Group', 'Total Modifier', 'Availability',
            'Current Fake Price (Rp)', 'Current Real Price (Rp)', 'Current Slash Price (%)', 'Current Slash Price (Rp)'
        ])
        
        # Use applicator name with first letter capitalized as OFD
        ofd_name = applicator.capitalize() if applicator else "Gofood"
        
        items_out['OFD'] = [ofd_name] * len(df_items)
        items_out['Outlet Name'] = df_items.get('Nama panjang', '')
        items_out['Outlet Short Name'] = df_items.get('Nama pendek (ShopeeFood)', '')
        items_out['Outlet Link'] = df_items.get('Link outlet', '')
        items_out['SID'] = df_items.get('Store ID', '')
        items_out['Category'] = df_items.get('Nama kategori', '')
        items_out['Item'] = df_items.get('Nama item', '')
        items_out['Photo Link'] = df_items.get('Link foto', '')
        items_out['Description'] = df_items.get('Deskripsi item', '')
        items_out['Total Sold'] = df_items.get('Jumlah terjual', '')
        items_out['Total Modifier Group'] = df_items.get('Jumlah modifier group', '')
        items_out['Total Modifier'] = df_items.get('Jumlah modifier', '')
        items_out['Availability'] = df_items.get('Ketersediaan item', '')

        if 'Harga item sebelum promo (harga coret)' in df_items:
            items_out['Current Fake Price (Rp)'] = df_items['Harga item sebelum promo (harga coret)']
        else:
            items_out['Current Fake Price (Rp)'] = 0

        if 'Harga item setelah promo (harga coret)' in df_items:
            items_out['Current Real Price (Rp)'] = df_items['Harga item setelah promo (harga coret)']
        else:
            items_out['Current Real Price (Rp)'] = items_out['Current Fake Price (Rp)']

        if 'Nominal atau persentase promo (harga coret)' in df_items:
            items_out['Current Slash Price (Rp)'] = df_items['Nominal atau persentase promo (harga coret)']
        else:
            items_out['Current Slash Price (Rp)'] = 0

        items_out['Current Slash Price (%)'] = 0
        fake = pd.to_numeric(items_out['Current Fake Price (Rp)'], errors='coerce').fillna(0)
        real = pd.to_numeric(items_out['Current Real Price (Rp)'], errors='coerce').fillna(0)
        mask = fake > 0
        items_out.loc[mask, 'Current Slash Price (%)'] = round((fake[mask] - real[mask]) / fake[mask] * 100)

        # Process Modifiers
        mods_out = pd.DataFrame(columns=[
            'OFD', 'Outlet Name', 'Outlet Short Name', 'Outlet Link', 'SID', 'Item', 'Modifier Group ID', 'Modifier Group',
            'Modifier ID', 'Modifier', 'Min', 'Max', 'Availability', 'Current Price (Rp)'
        ])
        mods_out['OFD'] = [ofd_name] * len(df_modifiers)
        mods_out['Outlet Name'] = df_modifiers.get('Nama panjang', '')
        mods_out['Outlet Short Name'] = df_modifiers.get('Nama pendek (ShopeeFood)', '')
        mods_out['Outlet Link'] = df_modifiers.get('Link outlet', '')
        mods_out['SID'] = df_modifiers.get('Store ID', '')
        mods_out['Item'] = df_modifiers.get('Nama item', '')
        mods_out['Modifier Group'] = df_modifiers.get('Nama modifier group', '')
        mods_out['Modifier'] = df_modifiers.get('Nama modifier', '')
        mods_out['Min'] = df_modifiers.get('Minimal', '')
        mods_out['Max'] = df_modifiers.get('Maksimal', '')
        mods_out['Availability'] = df_modifiers.get('Ketersediaan modifier', '')
        mods_out['Current Price (Rp)'] = df_modifiers.get('Harga modifier', '')

        wb = openpyxl.load_workbook(template_xlsx)

        ws_item = wb['Item']
        ws_item.delete_rows(2, ws_item.max_row)

        item_headers = [cell.value for cell in ws_item[1]]
        for r_idx, row in enumerate(items_out.to_dict('records'), start=2):
            for c_idx, header in enumerate(item_headers, start=1):
                if header in row and pd.notna(row[header]):
                    ws_item.cell(row=r_idx, column=c_idx, value=row[header])

        ws_mod = wb['Modifier']
        ws_mod.delete_rows(2, ws_mod.max_row)

        mod_headers = [cell.value for cell in ws_mod[1]]
        for r_idx, row in enumerate(mods_out.to_dict('records'), start=2):
            for c_idx, header in enumerate(mod_headers, start=1):
                if header in row and pd.notna(row[header]):
                    ws_mod.cell(row=r_idx, column=c_idx, value=row[header])

        safe_name = raw_outlet.replace('/', '_').replace('\\', '_')
        
        # Ambil Store ID dari data mentah agar nama file unik untuk setiap cabang
        store_id = str(df_items['Store ID'].iloc[0]).strip() if ('Store ID' in df_items and not df_items.empty) else ""
        
        if store_id:
            output_xlsx = os.path.join(output_dir, f"C5 - {safe_name} ({store_id}).xlsx")
        else:
            output_xlsx = os.path.join(output_dir, f"C5 - {safe_name}.xlsx")
        
        wb.save(output_xlsx)
        print(f"  [\033[92mC5 Generator\033[0m] Berhasil membuat file C5: {output_xlsx}")
        
        # Delete the raw Excel file as requested by user
        try:
            os.remove(raw_xlsx)
        except Exception as del_err:
            print(f"  [\033[93mC5 Generator\033[0m] Peringatan: Gagal menghapus file raw {raw_xlsx}: {del_err}")
            
        return output_xlsx
    except Exception as e:
        print(f"  [\033[91mC5 Generator\033[0m] Error saat membuat C5: {e}")
        return None
